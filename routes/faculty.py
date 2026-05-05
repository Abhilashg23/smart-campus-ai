from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from models import db, Event, ChatHistory, Attendance, User, SystemConfig, Subject, InternalMarks, Program, SubjectFeedback
from utils.decorators import faculty_required
from datetime import datetime
import io
from openpyxl import load_workbook, Workbook

faculty_bp = Blueprint('faculty', __name__)


@faculty_bp.route('/dashboard')
@login_required
@faculty_required
def dashboard():
    """Faculty dashboard"""
    from models import Attendance
    from datetime import datetime, date
    
    # Get upcoming events
    upcoming_events = Event.query.filter(
        Event.event_date >= datetime.utcnow()
    ).order_by(Event.event_date.asc()).limit(5).all()
    
    # Get statistics
    total_students = User.query.filter_by(role='Student', is_active=True).count()
    # Count subjects assigned to this faculty OR in the same department
    total_classes = Subject.query.join(Program).filter(
        (Subject.faculty_id == current_user.id) |
        (Program.department_id == current_user.department_id)
    ).count()
    
    # Get today's attendance count
    today = date.today()
    attendance_today = Attendance.query.filter(
        db.func.date(Attendance.timestamp) == today
    ).count()
    
    return render_template('faculty/dashboard.html',
                         faculty_name=current_user.full_name,
                         upcoming_events=upcoming_events,
                         total_students=total_students,
                         total_classes=total_classes,
                         attendance_today=attendance_today)


@faculty_bp.route('/events')
@login_required
@faculty_required
def events():
    """View all upcoming events"""
    upcoming_events = Event.query.filter(
        Event.event_date >= datetime.utcnow()
    ).order_by(Event.event_date.asc()).all()
    
    return render_template('faculty/events.html', events=upcoming_events)


@faculty_bp.route('/profile')
@login_required
@faculty_required
def profile():
    """Faculty profile page"""
    # Get events created by this faculty
    events_created = Event.query.filter_by(created_by=current_user.id).count()
    
    # Get students in same department
    students_count = User.query.filter_by(
        role='student',
        department_id=current_user.department_id
    ).count() if current_user.department_id else 0
    
    # Get chat interactions
    chat_count = ChatHistory.query.filter_by(user_id=current_user.id).count()
    
    return render_template('faculty/profile.html',
                         events_created=events_created,
                         students_count=students_count,
                         chat_count=chat_count)


@faculty_bp.route('/chat')
@login_required
@faculty_required
def chat():
    """Chatbot interface"""
    return render_template('faculty/chat.html')


@faculty_bp.route('/chat/history')
@login_required
@faculty_required
def chat_history():
    """View chat history"""
    history = ChatHistory.query.filter_by(
        user_id=current_user.id
    ).order_by(ChatHistory.timestamp.desc()).all()
    
    return render_template('faculty/chat_history.html', history=history)


@faculty_bp.route('/attendance')
@login_required
@faculty_required
def attendance():
    """View attendance records"""
    # Get students in the same department
    if current_user.department_id:
        students = db.session.query(User).filter_by(
            department_id=current_user.department_id,
            role='Student'
        ).all()
        
        attendance_records = Attendance.query.filter(
            Attendance.user_id.in_([s.id for s in students])
        ).order_by(Attendance.timestamp.desc()).limit(50).all()
    else:
        students = []
        attendance_records = []
    
    return render_template('faculty/attendance.html',
                         students=students,
                         attendance_records=attendance_records)


@faculty_bp.route('/anonymous-messages')
@login_required
@faculty_required
def anonymous_messages():
    """View anonymous messages from students"""
    return render_template('faculty/anonymous_message.html')


@faculty_bp.route('/api/monitor-attention', methods=['POST'])
@login_required
@faculty_required
def monitor_attention():
    """
    API endpoint for real-time attention monitoring
    Accepts video frame and returns attention analysis
    """
    from services.attention_monitoring import attention_service
    from models.attention_log import AttentionLog
    import cv2
    import numpy as np
    import uuid
    
    try:
        # Get image from request
        if 'image' not in request.files:
            return {'error': 'No image provided', 'success': False}, 400
        
        image_file = request.files['image']
        image_data = image_file.read()
        
        # Convert to numpy array
        nparr = np.frombuffer(image_data, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image is None:
            return {'error': 'Invalid image data', 'success': False}, 400
        
        # Analyze attention
        result = attention_service.analyze_attention(image)
        
        if not result['success']:
            return {'error': result.get('error', 'Analysis failed'), 'success': False}, 500
        
        # Get or create session ID from request
        session_id = request.form.get('session_id')
        if not session_id:
            session_id = str(uuid.uuid4())
        
        # Log to database (sample every 5 seconds to avoid spam)
        # Frontend should send session_id and only log periodically
        should_log = request.form.get('log', 'false').lower() == 'true'
        
        if should_log:
            attention_log = AttentionLog(
                faculty_id=current_user.id,
                session_id=session_id,
                total_students=result['total_faces'],
                focused_count=result['focused_count'],
                distracted_count=result['distracted_count'],
                alert_triggered=result['alert']
            )
            db.session.add(attention_log)
            db.session.commit()
        
        # Return analysis results
        return {
            'success': True,
            'session_id': session_id,
            'total_faces': result['total_faces'],
            'focused_count': result['focused_count'],
            'distracted_count': result['distracted_count'],
            'alert': result['alert'],
            'faces': result['faces']
        }
        
    except Exception as e:
        print(f"Error in monitor_attention: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'error': str(e), 'success': False}, 500


@faculty_bp.route('/marks')
@login_required
@faculty_required
def marks():
    """View assigned subjects and manage marks"""
    # ONLY show subjects directly assigned to this faculty member
    subjects = Subject.query.filter_by(faculty_id=current_user.id).order_by(
        Subject.program_id, Subject.semester
    ).all()
    
    # Annotate each subject with enrolled student count
    subject_student_counts = {}
    for subj in subjects:
        count = InternalMarks.query.filter_by(subject_id=subj.id).count()
        subject_student_counts[subj.id] = count
    
    is_locked = SystemConfig.get_value('marks_upload_locked', 'false') == 'true'
    
    return render_template('faculty/manage_marks.html', 
                           subjects=subjects,
                           subject_student_counts=subject_student_counts,
                           is_locked=is_locked)

@faculty_bp.route('/marks/<int:subject_id>/edit', methods=['GET', 'POST'])
@login_required
@faculty_required
def edit_marks(subject_id):
    """Manual entry/edit of internal marks"""
    subject = Subject.query.get_or_404(subject_id)
    
    # Only allow the assigned faculty to edit marks
    if subject.faculty_id != current_user.id:
        flash('You are not authorized to edit marks for this subject.', 'danger')
        return redirect(url_for('faculty.marks'))
        
    is_locked = SystemConfig.get_value('marks_upload_locked', 'false') == 'true'
    if is_locked:
        flash('Marks upload is currently locked by the administrator.', 'warning')
        return redirect(url_for('faculty.marks'))
        
    # Get students enrolled in this subject via InternalMarks records
    enrolled_records = InternalMarks.query.filter_by(subject_id=subject.id).all()
    student_ids = [r.student_id for r in enrolled_records]
    students = User.query.filter(User.id.in_(student_ids), User.role == 'Student').all()
    
    # Fallback: if no InternalMarks records exist, try program+semester match
    if not students:
        students = User.query.filter_by(
            role='Student',
            program_id=subject.program_id,
            semester=subject.semester
        ).all()
        # Auto-create InternalMarks enrolment records for these students
        for student in students:
            exists = InternalMarks.query.filter_by(
                student_id=student.id, subject_id=subject.id
            ).first()
            if not exists:
                db.session.add(InternalMarks(
                    student_id=student.id,
                    subject_id=subject.id,
                    marks=None,
                    max_marks=100.0
                ))
        db.session.commit()
    
    if request.method == 'POST':
        for student in students:
            marks_val = request.form.get(f'marks_{student.id}')
            if marks_val is not None and marks_val.strip() != '':
                try:
                    marks_val = float(marks_val)
                    record = InternalMarks.query.filter_by(
                        student_id=student.id, subject_id=subject.id
                    ).first()
                    if record:
                        record.marks = marks_val
                    else:
                        record = InternalMarks(
                            student_id=student.id, subject_id=subject.id,
                            marks=marks_val, max_marks=100.0
                        )
                        db.session.add(record)
                except ValueError:
                    pass
        db.session.commit()
        flash(f'Marks updated successfully for {subject.name}.', 'success')
        return redirect(url_for('faculty.marks'))
        
    # Pre-fetch existing marks
    marks_records = InternalMarks.query.filter_by(subject_id=subject.id).all()
    marks_dict = {m.student_id: m.marks for m in marks_records}
    
    return render_template('faculty/edit_marks.html', 
                           subject=subject, students=students, marks_dict=marks_dict)


@faculty_bp.route('/marks/<int:subject_id>/download')
@login_required
@faculty_required
def download_marks_template(subject_id):
    """Download excel template for marks entry"""
    subject = Subject.query.get_or_404(subject_id)
    
    if subject.faculty_id != current_user.id:
        flash('Unauthorized access', 'danger')
        return redirect(url_for('faculty.marks'))
        
    students = User.query.filter_by(role='Student', program_id=subject.program_id, semester=subject.semester).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Entry Template"
    
    # Headers
    ws.append(['Student ID', 'Student Name', 'Internal Marks (Max 30)'])
    
    marks_records = InternalMarks.query.filter_by(subject_id=subject.id).all()
    marks_dict = {m.student_id: m.marks for m in marks_records}
    
    for student in students:
        mark = marks_dict.get(student.id)
        mark_val = mark if mark is not None else ''
        # We need the user ID or registration ID. We'll use database ID to be safe, 
        # but display registration ID if available. Let's just put DB ID in a hidden way or first column.
        ws.append([student.id, student.full_name, mark_val])
        
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{subject.code}_marks_template.xlsx'
    )


@faculty_bp.route('/marks/<int:subject_id>/upload', methods=['POST'])
@login_required
@faculty_required
def upload_marks(subject_id):
    """Upload excel template for marks entry"""
    subject = Subject.query.get_or_404(subject_id)
    
    if subject.faculty_id != current_user.id:
        flash('Unauthorized access', 'danger')
        return redirect(url_for('faculty.marks'))
        
    is_locked = SystemConfig.get_value('marks_upload_locked', 'false') == 'true'
    if is_locked:
        flash('Marks upload is currently locked by the administrator.', 'warning')
        return redirect(url_for('faculty.marks'))
        
    if 'file' not in request.files:
        flash('No file provided', 'danger')
        return redirect(url_for('faculty.marks'))
        
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('faculty.marks'))
        
    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        
        updates_count = 0
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            student_id = int(row[0])
            marks_val = row[2]
            
            if marks_val is not None:
                try:
                    marks_val = float(marks_val)
                    record = InternalMarks.query.filter_by(student_id=student_id, subject_id=subject.id).first()
                    if record:
                        record.marks = marks_val
                    else:
                        record = InternalMarks(student_id=student_id, subject_id=subject.id, marks=marks_val)
                        db.session.add(record)
                    updates_count += 1
                except ValueError:
                    pass
                    
        db.session.commit()
        flash(f'Successfully updated marks for {updates_count} students.', 'success')
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'danger')
        
    return redirect(url_for('faculty.marks'))


@faculty_bp.route('/feedback')
@login_required
@faculty_required
def feedback():
    """Faculty view: see all student feedback for their subjects"""
    # Get subjects assigned to this faculty
    subjects = Subject.query.filter_by(faculty_id=current_user.id).order_by(
        Subject.program_id, Subject.semester
    ).all()

    feedback_data = []
    for subject in subjects:
        feedbacks = SubjectFeedback.query.filter_by(subject_id=subject.id).all()
        if not feedbacks:
            continue
        avg_rating = round(sum(f.rating for f in feedbacks) / len(feedbacks), 1)
        rating_counts = {i: sum(1 for f in feedbacks if f.rating == i) for i in range(1, 6)}
        feedback_data.append({
            'subject': subject,
            'feedbacks': feedbacks,
            'avg_rating': avg_rating,
            'count': len(feedbacks),
            'rating_counts': rating_counts,
        })

    # Overall average across all subjects
    all_feedbacks = SubjectFeedback.query.join(Subject).filter(
        Subject.faculty_id == current_user.id
    ).all()
    overall_avg = round(sum(f.rating for f in all_feedbacks) / len(all_feedbacks), 1) if all_feedbacks else 0

    return render_template('faculty/feedback.html',
                           feedback_data=feedback_data,
                           overall_avg=overall_avg,
                           total_responses=len(all_feedbacks))
